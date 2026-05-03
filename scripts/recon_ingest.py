#!/usr/bin/env python3
"""Full workbook recon ingest for SCN source."""

import hashlib
import json
import logging
import os
import re
import time
import unicodedata
import gc
import argparse
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

import psycopg2
from psycopg2.extras import execute_values
from openpyxl import load_workbook

try:
    import resource  # Unix-only
except ModuleNotFoundError:  # pragma: no cover - platform dependent
    resource = None

try:
    import psutil
except ModuleNotFoundError:  # pragma: no cover - optional dependency
    psutil = None

LOG_EVERY = int(os.getenv("RECON_LOG_EVERY", "10"))
COMMIT_EVERY = int(os.getenv("RECON_COMMIT_EVERY", "1000"))

FILES = {
    "content": "contentlicensing.xlsx",
    "pricing": "pricing.xlsx",
    "inventory": "inventory.xlsx",
}


def configure_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
    )


def log_event(event: str, **kwargs) -> None:
    """Write structured JSON logs so failures are easy to search in Render/local logs."""
    logging.info(json.dumps({"event": event, **kwargs}, default=str, ensure_ascii=False))


def normalize_key(value: str) -> str:
    txt = str(value or "").strip().lower()
    # Convert accented headers like Modèle to Modele before regex cleanup.
    txt = unicodedata.normalize("NFKD", txt).encode("ascii", "ignore").decode("ascii")
    txt = re.sub(r"[^a-z0-9]+", "_", txt)
    return re.sub(r"_+", "_", txt).strip("_")


def row_dict(headers, row):
    out = {}
    for idx, cell in enumerate(row):
        key = headers[idx] if idx < len(headers) else ""
        if key:
            out[key] = cell
    return out


def parse_decimal(value):
    if value in (None, ""):
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    txt = str(value).replace("$", "").replace(",", "").strip()
    if not txt:
        return None
    try:
        return Decimal(txt)
    except InvalidOperation:
        return None


def parse_ts(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    try:
        return datetime.fromisoformat(str(value).strip().replace("Z", "+00:00"))
    except ValueError:
        return None


def checksum(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        while True:
            chunk = f.read(1024 * 1024)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()


def get_conn():
    database_url = os.getenv("DATABASE_URL")
    if not database_url:
        raise RuntimeError("DATABASE_URL is required")
    return psycopg2.connect(database_url)


def memory_usage_mb() -> float:
    if resource is not None:
        # Linux ru_maxrss is in KB.
        return resource.getrusage(resource.RUSAGE_SELF).ru_maxrss / 1024
    if psutil is not None:
        return psutil.Process().memory_info().rss / (1024 * 1024)
    return 0.0


def log_memory(stage: str, elapsed_start: float | None = None) -> None:
    elapsed = ""
    if elapsed_start is not None:
        elapsed = f" elapsed={round(time.time() - elapsed_start, 2)}s"
    logging.info("%s memory_mb=%.2f%s", stage, memory_usage_mb(), elapsed)


def commit_and_log(conn, stage, rows):
    conn.commit()
    gc.collect()
    log_memory(stage)
    logging.info("%s committed rows=%s", stage, rows)

def log_every(stage: str, processed: int):
    if processed and processed % LOG_EVERY == 0:
        logging.info("%s processed=%s", stage, processed)


def ensure_source_and_locations(cur):
    cur.execute(
        """
        insert into probuy.primary_sources (code, name, is_active)
        values ('SCN', 'SCN International', true)
        on conflict (code) do nothing
        returning id
        """
    )
    row = cur.fetchone()
    if row:
        source_id = row[0]
    else:
        cur.execute("select id from probuy.primary_sources where code = 'SCN'")
        source_id = cur.fetchone()[0]

    locations = [
        ("SCN-CA", "SCN Canada Pricing", "National", "CA", {"source": "pricing.xlsx"}),
        ("MTL", "Montreal Distribution Centre", "QC", "CA", {"source": "inventory.xlsx", "sheet": "MTL"}),
        ("VAN", "Vancouver Distribution Centre", "BC", "CA", {"source": "inventory.xlsx", "sheet": "VAN"}),
        ("EDM", "Edmonton Distribution Centre", "AB", "CA", {"source": "inventory.xlsx", "sheet": "EDM"}),
    ]
    ids = {}
    for code, name, province, country, raw_data in locations:
        cur.execute(
            """
            insert into probuy.source_locations (source_id, code, name, province, country, is_active, raw_data)
            values (%s, %s, %s, %s, %s, true, %s::jsonb)
            on conflict (source_id, code) do nothing
            returning id
            """,
            (source_id, code, name, province, country, json.dumps(raw_data)),
        )
        row = cur.fetchone()
        if row:
            ids[code] = row[0]
        else:
            cur.execute("select id from probuy.source_locations where source_id = %s and code = %s", (source_id, code))
            ids[code] = cur.fetchone()[0]
    return source_id, ids


def insert_product(cur, source_id, key, model_no, brand, manufacturer, title, description, category, unit, raw_data):
    cur.execute(
        """
        insert into probuy.source_products
        (source_id, source_product_key, source_model_no, brand, manufacturer, product_title_en, description_en, category_en, unit_description_en, is_active, raw_data)
        values (%s,%s,%s,%s,%s,%s,%s,%s,%s,true,%s::jsonb)
        on conflict (source_id, source_product_key) do nothing
        returning id
        """,
        (source_id, key, model_no, brand, manufacturer, title, description, category, unit, json.dumps(raw_data, default=str)),
    )
    row = cur.fetchone()
    if row:
        return row[0]
    cur.execute("select id from probuy.source_products where source_id = %s and source_product_key = %s", (source_id, key))
    return cur.fetchone()[0]


def header_alias_map_content():
    return {
        "prod": "prod",
        "manufacturernumber": "manufacturer_number",
        "brand": "brand",
        "producttitle": "product_title",
        "categoryenglish": "category_english",
        "unitdescription": "unit_description",
        "date_last_modified": "date_last_modified",
    }


def header_alias_map_pricing():
    return {
        "model_no_no_modele": "model_no",
        "mfg_model_no_no_fab": "mfg_model_no",
        "fabricant": "manufacturer_fr",
        "manufacturer": "manufacturer",
        "english_description_description_anglais": "description_en",
        "french_description_description_francais": "description_fr",
        "list_price_prix_liste": "list_price",
        "distributor_cost_cout_distributeur": "dist_cost",
        "pricing_update_date_derniere_mise_a_jour_de_prix": "pricing_update_date",
        "category_level_1_english_categorie_niveau_1_anglais": "category_en",
        "unit_of_sale_unite_de_vente": "unit_of_sale",
    }


def to_alias_headers(headers, aliases):
    resolved = []
    for h in headers:
        nk = normalize_key(h)
        resolved.append(aliases.get(nk, nk))
    return resolved


def preview_row(row, limit=8):
    return list(row[:limit]) if row is not None else []


def log_sheet_start(file_type, path, wb, ws):
    log_event(
        "workbook_opened",
        file_type=file_type,
        path=str(path),
        sheets=wb.sheetnames,
        active_sheet=ws.title,
        max_row=ws.max_row,
        max_column=ws.max_column,
    )




def fetch_existing_product_ids(cur, source_id, keys):
    """Return {source_product_key: id} for a list/set of SCN product keys."""
    keys = [str(k).strip() for k in keys if str(k or "").strip()]
    if not keys:
        return {}
    cur.execute(
        """
        select source_product_key, id
        from probuy.source_products
        where source_id = %s
          and source_product_key = any(%s)
        """,
        (source_id, keys),
    )
    return {row[0]: row[1] for row in cur.fetchall()}


def preload_existing_price_model_keys(cur, source_id, loc_id):
    """Load model numbers that already have a price row for this source/location."""
    cur.execute(
        """
        select sp.source_product_key
        from probuy.source_product_prices spp
        join probuy.source_products sp on sp.id = spp.source_product_id
        where sp.source_id = %s
          and spp.location_id = %s
        """,
        (source_id, loc_id),
    )
    return {str(row[0]).strip() for row in cur.fetchall() if row[0]}


def preload_existing_inventory_model_keys(cur, source_id, location_id):
    """Load model numbers that already have an inventory row for this source/location."""
    cur.execute(
        """
        select sp.source_product_key
        from probuy.source_product_inventory spi
        join probuy.source_products sp on sp.id = spi.source_product_id
        where sp.source_id = %s
          and spi.location_id = %s
        """,
        (source_id, location_id),
    )
    return {str(row[0]).strip() for row in cur.fetchall() if row[0]}


def batch_insert_product_stubs(conn, cur, source_id, product_rows, stage):
    """
    Insert missing source_products in bulk and return {source_product_key: id}.
    product_rows is a list of dicts containing key/model/manufacturer/title/category/unit/raw_data.
    """
    if not product_rows:
        return {}

    keys = [r["key"] for r in product_rows]
    existing = fetch_existing_product_ids(cur, source_id, keys)
    missing = [r for r in product_rows if r["key"] not in existing]

    if missing:
        values = [
            (
                source_id,
                r["key"],
                r.get("model_no"),
                r.get("brand"),
                r.get("manufacturer"),
                r.get("title"),
                r.get("description"),
                r.get("category"),
                r.get("unit"),
                json.dumps(r.get("raw_data") or {}, default=str),
            )
            for r in missing
        ]
        execute_values(
            cur,
            """
            insert into probuy.source_products
            (source_id, source_product_key, source_model_no, brand, manufacturer, product_title_en, description_en, category_en, unit_description_en, is_active, raw_data)
            values %s
            on conflict (source_id, source_product_key) do nothing
            """,
            values,
            page_size=min(len(values), 5000),
        )
        log_event(stage + "_product_stub_inserted", attempted=len(values), inserted=cur.rowcount)

    # Query again so we get IDs for both existing and newly inserted rows.
    ids = fetch_existing_product_ids(cur, source_id, keys)
    return ids


def chunked(iterable, size):
    chunk = []
    for item in iterable:
        chunk.append(item)
        if len(chunk) >= size:
            yield chunk
            chunk = []
    if chunk:
        yield chunk


def backfill_product_images_from_source_products(conn, cur, source_id):
    """Populate product_images from known image keys in source_products.raw_data."""
    cur.execute(
        """
        with candidate_images as (
            select
                sp.id as source_product_id,
                coalesce(
                    nullif(sp.raw_data->>'image_main', ''),
                    nullif(sp.raw_data->>'image_file_name', ''),
                    nullif(sp.raw_data->>'image', ''),
                    nullif(sp.raw_data->>'image_url', ''),
                    nullif(sp.raw_data->'row'->>'image_main', ''),
                    nullif(sp.raw_data->'row'->>'image_file_name', ''),
                    nullif(sp.raw_data->'row'->>'image', ''),
                    nullif(sp.raw_data->'row'->>'image_url', '')
                ) as image_file_name
            from probuy.source_products sp
            where sp.source_id = %s
              and sp.is_active = true
        )
        insert into probuy.product_images (source_product_id, image_position, image_file_name, is_main_image, raw_data)
        select
            ci.source_product_id,
            1,
            ci.image_file_name,
            true,
            jsonb_build_object('source', 'recon_ingest_backfill', 'strategy', 'source_products.raw_data')
        from candidate_images ci
        where ci.image_file_name is not null
        on conflict (source_product_id, image_file_name) do update
        set
            is_main_image = true,
            image_position = coalesce(probuy.product_images.image_position, excluded.image_position),
            updated_at = now()
        """
        ,
        (source_id,),
    )
    upserted = max(cur.rowcount, 0)
    conn.commit()
    log_event("product_images_backfill_complete", source_id=str(source_id), upserted=upserted)
    return upserted

def ingest_content(conn, cur, source_id, counts, paths):
    stage_start = time.time()
    logging.info("file started: %s", paths["content"])
    wb = load_workbook(paths["content"], read_only=True, data_only=True)
    try:
        ws = wb.active
        ws.reset_dimensions()
        raw_headers = [str(c).strip() if c is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        headers = to_alias_headers(raw_headers, header_alias_map_content())
        logging.info("headers found content=%s", headers)
        pending = 0
        batch_seen = set()
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                rd = row_dict(headers, row)
                key = str(rd.get("prod", "")).strip()
                if not key or key in batch_seen:
                    continue
                batch_seen.add(key)

                product_id = insert_product(
                    cur, source_id, key, str(rd.get("manufacturer_number") or "").strip() or None, rd.get("brand"),
                    rd.get("brand"), rd.get("product_title"), None, rd.get("category_english"), rd.get("unit_description"),
                    {"source": "contentlicensing.xlsx", "row": rd},
                )
                counts["products"] += 1
                pending += 1
                if counts["products"] % LOG_EVERY == 0:
                    logging.info("content rows processed=%s", counts["products"])
                    log_memory("content-progress", stage_start)

                for i in range(1, 16):
                    an = rd.get(f"attributename{i}")
                    av = rd.get(f"attributevalue{i}")
                    if not an or av in (None, ""):
                        continue
                    canonical = normalize_key(str(an))
                    cur.execute(
                        """insert into probuy.attribute_definitions (canonical_name, display_name, data_type, unit, is_filterable, is_searchable)
                        values (%s,%s,'text',null,true,true)
                        on conflict (canonical_name) do nothing returning id""",
                        (canonical, str(an).strip()),
                    )
                    attr_row = cur.fetchone()
                    attr_id = attr_row[0] if attr_row else None
                    if attr_id is None:
                        cur.execute("select id from probuy.attribute_definitions where canonical_name = %s", (canonical,))
                        attr_id = cur.fetchone()[0]
                    cur.execute(
                        """insert into probuy.product_attribute_values (source_product_id, attribute_id, value_text, raw_data)
                        values (%s,%s,%s,%s::jsonb) on conflict (source_product_id, attribute_id) do nothing""",
                        (product_id, attr_id, str(av).strip(), json.dumps({"source": "contentlicensing.xlsx", "attribute_name": an, "attribute_value": av})),
                    )
                    counts["attributes"] += 1

                if pending >= COMMIT_EVERY:
                    commit_and_log(conn, "content", counts["products"])
                    logging.info("content rows inserted this batch=%s", pending)
                    batch_seen.clear()
                    pending = 0
            except Exception as exc:
                logging.exception("content row error row_number=%s error=%s", row_num, exc)
        if pending:
            commit_and_log(conn, "content-final", counts["products"])
            logging.info("content rows inserted this batch=%s", pending)
    finally:
        wb.close()


def ingest_pricing(conn, cur, source_id, loc_id, counts, paths):
    stage_start = time.time()
    batch_size = int(os.getenv("RECON_PRICE_BATCH_SIZE", str(COMMIT_EVERY)))
    skip_existing = os.getenv("RECON_SKIP_EXISTING", "true").lower() not in ("0", "false", "no")

    logging.info("file started: %s", paths["pricing"])
    wb = load_workbook(paths["pricing"], read_only=True, data_only=True)
    try:
        ws = wb.active
        ws.reset_dimensions()
        log_sheet_start("pricing", paths["pricing"], wb, ws)

        first_row = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        raw_headers = [str(c).strip() if c is not None else "" for c in first_row]
        normalized_headers = [normalize_key(h) for h in raw_headers]
        headers = to_alias_headers(raw_headers, header_alias_map_pricing())

        log_event(
            "pricing_headers_detected",
            raw_headers=raw_headers,
            normalized_headers=normalized_headers,
            resolved_headers=headers,
            has_model_no="model_no" in headers,
            has_list_price="list_price" in headers,
            has_dist_cost="dist_cost" in headers,
        )

        existing_price_models = preload_existing_price_model_keys(cur, source_id, loc_id) if skip_existing else set()
        log_event(
            "pricing_existing_preload_complete",
            skip_existing=skip_existing,
            existing_price_models=len(existing_price_models),
            batch_size=batch_size,
        )

        rows_seen = 0
        rows_ready = 0
        rows_inserted = 0
        missing_models = set()
        skipped_existing = 0
        skipped_missing_model = 0
        skipped_duplicate_in_file = 0
        sample_logged = 0
        seen_in_file = set()
        batch = []

        def flush_batch():
            nonlocal batch, rows_inserted
            if not batch:
                return
            product_ids = fetch_existing_product_ids(cur, source_id, [item["model_no"] for item in batch])

            values = []
            missing_product_id = 0
            for item in batch:
                rd = item["rd"]
                model_no = item["model_no"]
                product_id = product_ids.get(model_no)
                if not product_id:
                    missing_product_id += 1
                    missing_models.add(model_no)
                    continue
                values.append((
                    product_id,
                    loc_id,
                    model_no,
                    parse_decimal(rd.get("list_price")),
                    parse_decimal(rd.get("dist_cost")),
                    parse_ts(rd.get("pricing_update_date")),
                    parse_ts(rd.get("pricing_update_date")),
                    json.dumps({"source": "pricing.xlsx", "row": rd}, default=str),
                ))

            if values:
                execute_values(
                    cur,
                    """
                    insert into probuy.source_product_prices
                    (source_product_id, location_id, model_no, list_price, distributor_cost, currency_code, pricing_update_date, effective_at, raw_data)
                    values %s
                    on conflict (source_product_id, location_id) do nothing
                    """,
                    values,
                    template="(%s,%s,%s,%s,%s,'CAD',%s,%s,%s::jsonb)",
                    page_size=min(len(values), 5000),
                )
                inserted_now = max(cur.rowcount, 0)
                rows_inserted += inserted_now
                counts["prices"] += inserted_now

            conn.commit()
            for item in batch:
                existing_price_models.add(item["model_no"])

            log_event(
                "pricing_batch_complete",
                batch_rows=len(batch),
                db_values=len(values),
                inserted_now=inserted_now if values else 0,
                total_inserted=rows_inserted,
                missing_product_id=missing_product_id,
                rows_seen=rows_seen,
                skipped_existing=skipped_existing,
                skipped_missing_model=skipped_missing_model,
                skipped_duplicate_in_file=skipped_duplicate_in_file,
            )
            log_memory("pricing-batch", stage_start)
            batch = []
            gc.collect()

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            rows_seen += 1
            rd = row_dict(headers, row)
            model_no = str(rd.get("model_no", "")).strip()

            if rows_seen <= 3:
                log_event("pricing_sample_row", row_num=row_num, row_preview=preview_row(row), mapped_preview={k: rd.get(k) for k in headers[:8]})

            if not model_no:
                skipped_missing_model += 1
                if sample_logged < 5:
                    log_event("pricing_row_skipped", row_num=row_num, reason="missing_model_no", row_preview=preview_row(row), mapped_keys=list(rd.keys())[:20])
                    sample_logged += 1
                continue

            if model_no in seen_in_file:
                skipped_duplicate_in_file += 1
                continue
            seen_in_file.add(model_no)

            if skip_existing and model_no in existing_price_models:
                skipped_existing += 1
                continue

            batch.append({"row_num": row_num, "model_no": model_no, "rd": rd})
            rows_ready += 1

            if rows_seen % LOG_EVERY == 0:
                log_event(
                    "pricing_progress",
                    rows_seen=rows_seen,
                    rows_ready=rows_ready,
                    rows_inserted=rows_inserted,
                    skipped_existing=skipped_existing,
                    skipped_missing_model=skipped_missing_model,
                    skipped_duplicate_in_file=skipped_duplicate_in_file,
                    current_batch=len(batch),
                )

            if len(batch) >= batch_size:
                flush_batch()

        flush_batch()

        log_event(
            "pricing_parse_summary",
            rows_seen=rows_seen,
            rows_ready=rows_ready,
            rows_inserted=rows_inserted,
            skipped_existing=skipped_existing,
            skipped_missing_model=skipped_missing_model,
            skipped_duplicate_in_file=skipped_duplicate_in_file,
            elapsed_seconds=round(time.time() - stage_start, 2),
        )
        if missing_models:
            with open("notfound.txt", "a", encoding="utf-8") as fh:
                for model_no in sorted(missing_models):
                    fh.write(f"pricing\t{model_no}\n")
    finally:
        wb.close()

def ingest_inventory(conn, cur, source_id, loc_ids, counts, paths):
    stage_start = time.time()
    batch_size = int(os.getenv("RECON_INVENTORY_BATCH_SIZE", str(COMMIT_EVERY)))
    skip_existing = os.getenv("RECON_SKIP_EXISTING", "true").lower() not in ("0", "false", "no")

    logging.info("file started: %s", paths["inventory"])
    wb = load_workbook(paths["inventory"], read_only=True, data_only=True)

    total_seen = 0
    total_ready = 0
    total_inserted = 0
    total_skipped_existing = 0
    total_skipped_missing_model = 0
    total_skipped_duplicate = 0
    missing_models = set()

    try:
        log_event("inventory_workbook_opened", path=str(paths["inventory"]), sheets=wb.sheetnames, batch_size=batch_size, skip_existing=skip_existing)

        for sheet_name in ["MTL", "VAN", "EDM"]:
            if sheet_name not in wb.sheetnames:
                log_event("inventory_sheet_skipped", sheet=sheet_name, reason="sheet_missing", available_sheets=wb.sheetnames)
                continue

            location_id = loc_ids[sheet_name]
            existing_inventory_models = preload_existing_inventory_model_keys(cur, source_id, location_id) if skip_existing else set()
            log_event(
                "inventory_existing_preload_complete",
                sheet=sheet_name,
                skip_existing=skip_existing,
                existing_inventory_models=len(existing_inventory_models),
            )

            ws = wb[sheet_name]
            ws.reset_dimensions()

            first_row = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
            raw_headers = [str(c).strip() if c is not None else "" for c in first_row]
            headers = [normalize_key(h) for h in raw_headers]

            log_event(
                "inventory_headers_detected",
                sheet=sheet_name,
                raw_headers=raw_headers,
                normalized_headers=headers,
                has_model_no_header="model_no_no_modele" in headers,
                has_quantity_header="quantity_available_quantite_disponible" in headers,
                max_row=ws.max_row,
                max_column=ws.max_column,
            )

            rows_seen = 0
            rows_ready = 0
            rows_inserted = 0
            skipped_existing = 0
            skipped_missing_model = 0
            skipped_duplicate_in_sheet = 0
            sample_logged = 0
            seen_in_sheet = set()
            batch = []

        def flush_batch():
            nonlocal batch, rows_inserted, total_inserted
            if not batch:
                return
            product_ids = fetch_existing_product_ids(cur, source_id, [item["model_no"] for item in batch])

            values = []
            missing_product_id = 0
            for item in batch:
                rd = item["rd"]
                model_no = item["model_no"]
                product_id = product_ids.get(model_no)
                if not product_id:
                    missing_product_id += 1
                    missing_models.add(model_no)
                    continue
                values.append((
                    product_id,
                    location_id,
                    model_no,
                    rd.get("status_etat_des_stocks") or rd.get("stock_status_etat_des_stocks"),
                    parse_decimal(rd.get("quantity_available_quantite_disponible")) or Decimal("0"),
                    parse_ts(rd.get("inventory_update_date_date_de_mise_a_jour_de_l_inventaire")),
                    json.dumps({"source": "inventory.xlsx", "sheet": sheet_name, "row": rd}, default=str),
                ))

            inserted_now = 0
            if values:
                execute_values(
                    cur,
                    """
                    insert into probuy.source_product_inventory
                    (source_product_id, location_id, model_no, stock_status, quantity_available, inventory_update_date, raw_data)
                    values %s
                    on conflict (source_product_id, location_id) do nothing
                    """,
                    values,
                    template="(%s,%s,%s,%s,%s,%s,%s::jsonb)",
                    page_size=min(len(values), 5000),
                )
                inserted_now = max(cur.rowcount, 0)
                rows_inserted += inserted_now
                total_inserted += inserted_now
                counts["inventory"] += inserted_now

            conn.commit()
            for item in batch:
                existing_inventory_models.add(item["model_no"])

            log_event(
                "inventory_batch_complete",
                sheet=sheet_name,
                batch_rows=len(batch),
                db_values=len(values),
                inserted_now=inserted_now,
                sheet_inserted=rows_inserted,
                total_inserted=total_inserted,
                missing_product_id=missing_product_id,
                sheet_rows_seen=rows_seen,
                skipped_existing=skipped_existing,
                skipped_missing_model=skipped_missing_model,
                skipped_duplicate_in_sheet=skipped_duplicate_in_sheet,
            )
            log_memory(f"inventory-batch-{sheet_name}", stage_start)
            batch = []
            gc.collect()

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                rows_seen += 1
                total_seen += 1
                rd = row_dict(headers, row)
                model_no = str(rd.get("model_no_no_modele", "")).strip()

                if rows_seen <= 3:
                    log_event("inventory_sample_row", sheet=sheet_name, row_num=row_num, row_preview=preview_row(row), mapped_preview={k: rd.get(k) for k in headers[:8]})

                if not model_no:
                    skipped_missing_model += 1
                    total_skipped_missing_model += 1
                    if sample_logged < 5:
                        log_event("inventory_row_skipped", sheet=sheet_name, row_num=row_num, reason="missing_model_no", row_preview=preview_row(row), mapped_keys=list(rd.keys())[:20])
                        sample_logged += 1
                    continue

                if model_no in seen_in_sheet:
                    skipped_duplicate_in_sheet += 1
                    total_skipped_duplicate += 1
                    continue
                seen_in_sheet.add(model_no)

                if skip_existing and model_no in existing_inventory_models:
                    skipped_existing += 1
                    total_skipped_existing += 1
                    continue

                batch.append({"row_num": row_num, "model_no": model_no, "rd": rd})
                rows_ready += 1
                total_ready += 1

                if rows_seen % LOG_EVERY == 0:
                    log_event(
                        "inventory_progress",
                        sheet=sheet_name,
                        sheet_rows_seen=rows_seen,
                        sheet_rows_ready=rows_ready,
                        sheet_inserted=rows_inserted,
                        skipped_existing=skipped_existing,
                        skipped_missing_model=skipped_missing_model,
                        skipped_duplicate_in_sheet=skipped_duplicate_in_sheet,
                        current_batch=len(batch),
                    )

                if len(batch) >= batch_size:
                    flush_batch()

            flush_batch()

            log_event(
                "inventory_sheet_summary",
                sheet=sheet_name,
                rows_seen=rows_seen,
                rows_ready=rows_ready,
                rows_inserted=rows_inserted,
                skipped_existing=skipped_existing,
                skipped_missing_model=skipped_missing_model,
                skipped_duplicate_in_sheet=skipped_duplicate_in_sheet,
            )

        log_event(
            "inventory_parse_summary",
            rows_seen=total_seen,
            rows_ready=total_ready,
            rows_inserted=total_inserted,
            skipped_existing=total_skipped_existing,
            skipped_missing_model=total_skipped_missing_model,
            skipped_duplicate_in_sheet=total_skipped_duplicate,
            elapsed_seconds=round(time.time() - stage_start, 2),
        )
        if missing_models:
            with open("notfound.txt", "a", encoding="utf-8") as fh:
                for model_no in sorted(missing_models):
                    fh.write(f"inventory\t{model_no}\n")
    finally:
        wb.close()

def main():
    parser = argparse.ArgumentParser(description="Full workbook recon ingest for SCN source.")
    parser.add_argument(
        "--start-at",
        choices=["content", "pricing", "inventory"],
        default="content",
        help="Start ingest from this phase (default: content).",
    )
    parser.add_argument(
        "--input-dir",
        default="../input/data",
        help="Folder containing contentlicensing.xlsx, pricing.xlsx, and inventory.xlsx.",
    )
    parser.add_argument(
        "--no-skip-existing",
        action="store_true",
        help="Do not preload existing price/inventory rows. Useful only when testing conflict behavior.",
    )
    parser.add_argument(
        "--images-only",
        action="store_true",
        help="Only backfill product_images from source_products.raw_data and exit.",
    )
    args = parser.parse_args()

    configure_logging()
    start = time.time()
    Path("notfound.txt").write_text("", encoding="utf-8")

    if args.no_skip_existing:
        os.environ["RECON_SKIP_EXISTING"] = "false"

    input_dir = Path(args.input_dir)
    paths = {k: input_dir / v for k, v in FILES.items()}
    if not args.images_only:
        missing = [str(p) for p in paths.values() if not p.exists()]
        if missing:
            raise FileNotFoundError(f"Missing required files: {', '.join(missing)}")

    info = {}
    if not args.images_only:
        info = {k: {"path": str(v), "sha256": checksum(v)} for k, v in paths.items()}
    counts = {"products": 0, "attributes": 0, "prices": 0, "inventory": 0}

    logging.info("Starting full recon ingest with streaming workbooks")
    logging.info("Purge step disabled for recon ingest")
    logging.info("Config log_every=%s commit_every=%s", LOG_EVERY, COMMIT_EVERY)

    with get_conn() as conn:
        with conn.cursor() as cur:
            source_id, loc_ids = ensure_source_and_locations(cur)
            conn.commit()

            phase_order = ["content", "pricing", "inventory"]
            start_index = phase_order.index(args.start_at)
            phases_to_run = phase_order[start_index:]
            logging.info("Running phases=%s", phases_to_run)

            if args.images_only:
                backfill_product_images_from_source_products(conn, cur, source_id)
            else:
                if "content" in phases_to_run:
                    ingest_content(conn, cur, source_id, counts, paths)
                if "pricing" in phases_to_run:
                    ingest_pricing(conn, cur, source_id, loc_ids["SCN-CA"], counts, paths)
                if "inventory" in phases_to_run:
                    ingest_inventory(conn, cur, source_id, loc_ids, counts, paths)
                backfill_product_images_from_source_products(conn, cur, source_id)

    elapsed = round(time.time() - start, 3)
    logging.info("Full recon ingest complete in %ss", elapsed)
    print(json.dumps({"event": "ingestfullrecon_complete", "elapsed_seconds": elapsed, "counts": counts, "files": info}, ensure_ascii=False))


if __name__ == "__main__":
    main()
